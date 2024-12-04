import asyncio
import csv
import json
import yaml
import time
import argparse
import nest_asyncio
from tqdm import tqdm
from datetime import datetime
from conva_ai import AsyncConvaAI
from openpyxl import load_workbook
from conva_ai.response import ConvaAIResponse
# import MS_consolidated_compare


nest_asyncio.apply()
current_date = datetime.now().strftime("%d")

def load_config(env, date):
    # with open('../polyglot/tools/testing/loadtest/MS_config_SDK_Python.yaml', 'r') as file:
    with open('/Users/divyac/Documents/Notebooks/MS_config_SDK_Python.yaml', 'r') as file:
        config = yaml.safe_load(file)
    
    if env in config['configurations']:
        env_config = config['configurations'][env]
        env_config['output_file_path'] = env_config['output_file_path'].replace("+current_date+", date)
        return env_config
    else:
        raise ValueError("Invalid environment specified")
    

def extract_properties(dict: ConvaAIResponse):
  request_id = ""
  try:
    request_id = dict.request_id
    message = dict.message
    related_queries = dict.related_queries
    tool_name = dict.tool_name
    reason = dict.reason
    steps =  dict.parameters.get('steps', [])
    api_response =  dict.parameters.get('code_sample', {})
    citations = dict.parameters.get('citations', [])
    parameters = dict.parameters
    return request_id, message, related_queries, tool_name, reason, steps, api_response, citations, parameters

  except json.JSONDecodeError:
    return "", "", "", "", "", [], {}, [], {}


def generate_ouput(output_file_path, input_file_path, input_sheet_name):
    with open(output_file_path, "w") as outfile:
        wb = load_workbook(input_file_path)
        writer = csv.writer(outfile)

        if input_sheet_name in wb.sheetnames:
            print(input_sheet_name)
            sheet = wb[input_sheet_name]
            ws_index = wb.index(sheet)
            wb.active = ws_index
            ws = wb.active
            iter_rows = ws.iter_rows(min_row=2, values_only=True)
            writer.writerow(
                ["Assistant_name", "request_id", "input_query", "message", "related_queries", "tool_name", "reason", "steps", "api_response", "citations", "parameters", "json"]
            )

        for row in iter_rows:
            assis_id = row[0]
            assistant_version = row[1]
            API_KEY = row[2]
            sentence = row[3]
            ass_name = row[5]
            print("\n\n", "AssistantID is :", assis_id)
            client = AsyncConvaAI(assistant_id=assis_id, assistant_version=assistant_version, api_key=API_KEY, host=HOST)
            print("Sentence:", sentence)
            final_response = asyncio.run(client.invoke_capability(sentence, stream=False, llm_key = "openai-gpt-4o-mini-2024-07-18", timeout=500))
            print("final response:",final_response)
            request_id, message, related_queries, tool_name, reason, steps, api_response, citations, parameters= extract_properties(final_response)
            row = [ass_name, request_id, sentence, message, related_queries, tool_name, reason, steps, api_response, citations, parameters, final_response]
            writer.writerow(row)
            time.sleep(1)
            
if __name__ == "__main__":
    current_date = datetime.now().strftime("%d")
    parser = argparse.ArgumentParser(description='Load configuration based on the environment.')
    parser.add_argument('environment', type=str, help='The environment to load configuration for (stage_amazon, prod_amazon, dev_weechef)')
    
    args = parser.parse_args()

    try:
        config = load_config(args.environment, current_date)
        # Assigning output_file_path from the loaded configuration
        output_file_path = config['output_file_path']
        input_file_path = config['input_file_path']
        input_sheet_name = config['input_sheet_name']
        HOST = config['HOST']
        print("output:", output_file_path)
        print("Input:", input_sheet_name)
        print("sheet:", output_file_path)
        print("Host:", HOST)
        generate_ouput(output_file_path, input_file_path, input_sheet_name)
    except ValueError as e:
        print(e)
#  To run
#  /Users/divyac/Documents/Notebooks/my_env/bin/python /Users/divyac/Documents/Notebooks/my_env/MS_consolidated_prompt_testing.py stage_consolidated_prompt
