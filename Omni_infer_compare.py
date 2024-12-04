import csv
import yaml
import argparse
from datetime import datetime, timezone
import json
import base64
import requests
import hashlib
import uuid
import hmac
from openpyxl import load_workbook
from cryptography.fernet import Fernet
import sseclient
from tqdm import tqdm
from time import perf_counter_ns
import pandas as pd
import json
import csv
from sklearn.metrics import f1_score
import torch
from transformers import AutoTokenizer, AutoModel
from sklearn.metrics.pairwise import cosine_similarity
from sklearn.utils.multiclass import unique_labels
from tqdm import tqdm 
import numpy as np

device_ids = [
    "INT-TEST-1234",
    "INT-TEST-1235",
    "INT-TEST-1236",
    "INT-TEST-1237",
    "INT-TEST-1238",
    "INT-TEST-1239",
    "INT-TEST-1240",
    "INT-TEST-1241",
    "INT-TEST-1243",
    "INT-TEST-1244"
]
device_id_counter = 0

def load_config(env, date):
    with open('../polyglot/tools/testing/loadtest/configurations.yaml', 'r') as file:
        config = yaml.safe_load(file)

    current_date = datetime.now().strftime("%d")
    
    if env in config['configurations']:
        env_config = config['configurations'][env]
        env_config['output_file_path'] = env_config['output_file_path'].replace("+current_date+", date)
        env_config['file2'] = env_config['file2'].replace("+current_date+", date)
        env_config['file3'] = env_config['file3'].replace("+current_date+", date)
        return env_config
    else:
        raise ValueError("Invalid environment specified")


current_date = datetime.now().strftime("%d")
input_csv_encoding = 'utf-8'
output_csv_encoding = 'utf-8'

if __name__ == "__main__":
    current_date = datetime.now().strftime("%d")
    parser = argparse.ArgumentParser(description='Load configuration based on the environment.')
    parser.add_argument('environment', type=str, help='The environment to load configuration for (stage, prod, mylo)')
    
    args = parser.parse_args()
    
    try:
        config = load_config(args.environment, current_date)
        # Assigning output_file_path from the loaded configuration
        output_file_path = config['output_file_path']
        input_file_path = config['input_file_path']
        input_sheet_name = config['input_sheet_name']
        HOST = config['HOST']
        API_KEY = config['API_KEY']
        file1 = config['file1']
        file2 = config['file2']
        file3 = config['file3']
        assistant_version = config['assistant_version']
        # print(output_file_path)
        # print(input_file_path)
        # print(input_sheet_name)
        # print(HOST)
        # print(API_KEY)
        # print(file1)
        # print(file2)
        # print(file3)
    except ValueError as e:
        print(e)

def generate_security_token(request_key, request_timestamp) -> str:
    salt = bytes("7j7efiQmVrCs43NA8wrzmcH2xCqWNoE6_C2rgqQSVTQ=".encode())
    fernet = Fernet(salt)
    return fernet.encrypt(f"{request_key}:{request_timestamp}".encode()).decode()


def generate_auth_hash(
    api_key, request_path, request_method, request_key, request_timestamp
) -> str:
    signing_key = base64.b64encode(
        f"{request_path}:{request_method}:{request_key}:{request_timestamp}".encode()
    )
    digest = hmac.new(signing_key, api_key.encode(), digestmod=hashlib.sha256).digest()
    computed_hmac = base64.b64encode(digest)
    return computed_hmac.decode()


def generate_auth_creds(
    api_key,
    request_path,
    request_method="POST",
) -> tuple[str, str]:
    request_key = uuid.uuid4().hex
    request_timestamp = (datetime.now(tz=timezone.utc)).timestamp()
    apikey_hash = generate_auth_hash(
        api_key, request_path, request_method, request_key, request_timestamp
    )
    token = generate_security_token(request_key, request_timestamp)
    return apikey_hash, token


def _create_request_path(assistant_id, request_type):
    return f"/v1/assistants/{assistant_id}/{request_type}"

def extract_properties(json_str):
  try:
    # Check if the input is a valid JSON string
    if isinstance(json_str, str):
      json_obj = json.loads(json_str)

      app_name = json_obj.get("app_name", "")
      app_action = json_obj.get("app_action", "")
      category = json_obj.get("category", "")
      app_name = json_obj.get("app_name", "")
      hints = json_obj.get("hints", [])

      suggestions  = []
      nav_url = ""
      # Extract text values from the suggestions
      for suggest in json_obj.get("suggestions", []):
        text = suggest.get("display", {}).get("text", "")
        if text:
            suggestions.append(text)

      message = json_obj.get("message", "")
      search_term = json_obj.get("parameters", {}).get("search_term", "")
      tool_name = json_obj.get("tool_name", "")
      filter_pairs = []
      nav_url = json_obj.get("parameters", {}).get("navigation_target_url", "")

      # Extract filters (if they exist)
      if "parameters" in json_obj and "filters" in json_obj["parameters"]:
        filters = json_obj["parameters"]["filters"]
        for filter_info in filters:
          # Attempt to extract 'key' and 'value' from various locations:
          if "key" in filter_info and "parameters" in filter_info and "value" in filter_info["parameters"]:
            filter_key = filter_info["key"]
            filter_value = filter_info["parameters"]["value"]
          elif "filter_key" in filter_info and "filter_value" in filter_info:  # Check for alternative key names
            filter_key = filter_info["filter_key"]
            filter_value = filter_info["filter_value"]
          else:
            print("Warning: Unexpected filter format:", filter_info) 
            continue  # Skip to the next filter 
        
          filter_pairs.append((filter_key, filter_value))

      return app_name, app_action, category, hints, suggestions, message, search_term, tool_name, filter_pairs, nav_url

    else:
      return "", "", "", "", [], "", "", "", [], ""
  except json.JSONDecodeError:
    return "", "", "", "", [],  "", "", "", [], ""

def _do_text2action_request(body, assistant_id):
    request_type = "text2action"
    uri = _create_request_path(assistant_id=assistant_id, request_type=request_type)
    formatted_url = f"{HOST}{uri}"
    apikey_hash, token = generate_auth_creds(
        API_KEY,
        uri,
    )
    request_headers = {
        "Authorization": apikey_hash,
        "X-Slang-Security-Token": token,
        "Content-Type": "application/json",
    }
    start_time = perf_counter_ns()  # Start the timer
    response = requests.post(
        formatted_url,
        stream=True,
        json=body,
        headers=request_headers,
    )
    client = sseclient.SSEClient(response)
    time_to_first_token = 0
    latency = None

    tmp = []
    for event in client.events():
        if time_to_first_token == 0:
            time_to_first_token = (perf_counter_ns() - start_time) / 1e6 # Time to first token
        tmp.append(event.data)

    latency = (perf_counter_ns() - start_time) / 1e6  # Total latency
    result = tmp[-1] if tmp else None
    # print(result)
    return result, time_to_first_token, latency


with open(output_file_path, "w") as outfile:
    wb = load_workbook(input_file_path)
    writer = csv.writer(outfile)

    if input_sheet_name in wb.sheetnames:
        sheet = wb[input_sheet_name]
        ws_index = wb.index(sheet)
        wb.active = ws_index
        ws = wb.active
        iter_rows = ws.iter_rows(min_row=2, values_only=True)
        writer.writerow(
            ["assistant_id", "query", "app_name", "app_action", "category", "hints",  "suggestions", "message", "search_term", "tool_name", "filter_pairs", "navigation_url"]
        )

        for row in tqdm(iter_rows, desc="Processing rows", unit="row", dynamic_ncols=True):
            print(row[0])
            assistant_id = row[0]
            sentence = row[3]
            request_id = str(uuid.uuid4())
            expected_response = row[4]
            device_id = device_ids[device_id_counter % len(device_ids)]
            device_id_counter += 1
            
            params = {
                "type": "text2action",  # text2Search
                "assistant_id": assistant_id,
                "assistant_version": assistant_version,
                "request_id": request_id,
                "device_id": device_id,
                "tool_config": {
                    "multi_step_tool": {"llm": "azure-gpt-3.5-turbo"},
                    "AssistedShopping": {"llm": "azure-gpt-3.5-turbo"},
                    "single_step_tool": {"llm": "azure-gpt-3.5-turbo"},
                    "router": {"llm": "azure-gpt-3.5-turbo"}
                },
                "domain_name": "retail",
                "input_query": sentence,
                "app_context": {},
            }
            result, time_to_first_token, latency = _do_text2action_request(params, assistant_id)
            app_name, app_action, category, hints,  suggestions, message, search_term, tool_name, filter_pairs, nav_url = extract_properties(str(result))
            writer.writerow(
                [
                    assistant_id,
                    sentence,
                    app_name, 
                    app_action, 
                    category, 
                    hints, 
                    suggestions, 
                    message, 
                    search_term, 
                    tool_name, 
                    filter_pairs, 
                    nav_url
                ]
            )


tokenizer = AutoTokenizer.from_pretrained('BAAI/bge-large-en-v1.5', truncation=True)
model = AutoModel.from_pretrained('BAAI/bge-large-en-v1.5')

def compute_cosine_similarity(true_values, predicted_values):
    
  true_encoding = tokenizer(true_values, return_tensors='pt', truncation=True)
  predicted_encoding = tokenizer(predicted_values, return_tensors='pt', truncation=True)

  with torch.no_grad():
    true_output = model(**true_encoding).last_hidden_state.mean(dim=1)
    predicted_output = model(**predicted_encoding).last_hidden_state.mean(dim=1)

  similarity = cosine_similarity(true_output, predicted_output)
  return similarity[0][0]

def compare_csv_files(file1, file2, output_file):
    # Read data from the first CSV file into a list
    with open(file1, 'r', newline='') as csv_file1:
        csv_reader1 = csv.reader(csv_file1)
        data1 = list(csv_reader1)
    
    # Read data from the second CSV file into a list
    with open(file2, 'r', newline='') as csv_file2:
        csv_reader2 = csv.reader(csv_file2)
        data2 = list(csv_reader2)
    
    columns_to_compare = [0,1,2,3,4,5,6,7,8,9,10,11]
    # Assuming both data1 and data2 are lists of lists where each inner list represents a row
    mismatched_columns = []
    for index in range(min(len(data1), len(data2))):
        row3 = []
        row1 = data1[index]
        row2 = data2[index]
        App_name_score = App_action_score = Category_score = Suggestions_score = Message_score = Search_term_score = Tool_name_score = Filter_pairs_score = url_score = 0
        for col_index in columns_to_compare:
                flag=0
                list1 = [i.strip("[]' ").lower() for i in row1[col_index].split(',')]
                if ',' in row2[col_index]:
                    list2 = [i.strip("[]' ").lower() for i in row2[col_index].split(',')]
                else:
                    list2 = [row2[col_index]]

                if col_index == 2:
                    for i in list2:
                        for j in list1:
                            if j.lower() in i.lower():
                                flag=1
                                App_name_score = 1
                                break

                elif col_index == 3:
                    for i in list2:
                        for j in list1:
                            if j.lower() in i.lower():
                                flag=1
                                App_action_score = 1
                                break

                elif col_index == 4:
                    for i in list2:
                        for j in list1:
                            if j.lower() in i.lower():
                                flag=1
                                Category_score = 1
                                break

                elif col_index == 6: #suggestions
                    # flag = 1
                    for i in list2:
                        for j in list1:
                            if j.lower() in i.lower():
                                flag=1
                                Suggestions_score = 1
                                break

                elif col_index == 7: #cousine similaritu
                    for i in list2:
                        for j in list1:
                            Message_score = compute_cosine_similarity(i.lower(), j.lower())
                            if Message_score >= 0.5 and Message_score <= 2.0:
                                flag=1
                                break
                    
                elif col_index == 8:   #search term
                    unwanted_list = ['for', 'with']
                    # flag = 1
                    for i in list2:
                        for j in list1:
                            if i.lower() == "apple electronics" or i.lower() == "apple electronic" or i.lower() in unwanted_list:
                                break
                            if j.lower() in i.lower():
                                flag=1
                                Search_term_score = 1
                                break

                elif col_index == 9:
                    for i in list2:
                        for j in list1:
                            if j.lower() in i.lower():
                                flag=1
                                Tool_name_score = 1
                                break
                        
                elif col_index == 10:
                    for i in list2:
                        for j in list1:
                            if j.lower() in i.lower():
                                flag=1
                                Filter_pairs_score = 1
                                break


                elif col_index == 11:
                    for i in list2:
                        for j in list1:
                            if j.lower() in i.lower():
                                flag=1
                                url_score = 1
                                break

                else:
                    for i in list2:
                        for j in list1:
                            if j.lower() in i.lower():
                                flag=1
                                break
                if flag == 0:
                    row3 = [row2[0].split(','), row2[1].split(','), row1[2], row2[2].split(','), App_name_score, row1[3], row2[3].split(','), App_action_score, row1[4], row2[4].split(','), Category_score, row2[5], row1[6], row2[6].split(','), Suggestions_score, row1[7], row2[7].split(','), Message_score, row1[8], row2[8].split(','), Search_term_score, row1[9], row2[9].split(','), Tool_name_score, row1[10], row2[10], Filter_pairs_score, row1[11], row2[11], url_score]
                    mismatched_columns.append(row3)
                    print("Actual row : ", [row2[1]], " ", list2)
                    print("Expected row : ", row1[1], " ", list1)
                    break
            

    # Write the mismatched rows to the output CSV file
    with open(output_file, 'w', newline='') as output_csv:
        csv_writer = csv.writer(output_csv)
        csv_writer.writerow(
            ["assistant_id", "query", "Expected_app_name", "app_name", "App_name_score", "Expected_app_action", "app_action", "App_action_score", "Expected_category", "category", "Category_score", "hints",  "Expected_suggestions", "suggestions", "Suggestions_score",  "Expected_message", "message", "Message_score", "Expected_search_term", "search_term", "Search_term_score", "Expected_tool_name", "tool_name", "Tool_name_score", "Expected_filter_pairs", "filter_pairs", "Filter_pairs_score", "Expected_url", "navigation_url", "url_score"]
        )
        csv_writer.writerows(mismatched_columns)


compare_csv_files(file1, file2, file3)

#  To run
#  ../polyglot/.venv/bin/python ../polyglot/tools/testing/loadtest/Omni_infer_compare.py stage  