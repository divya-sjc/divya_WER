import asyncio
import csv
import json
import yaml
import argparse
import nest_asyncio
from tqdm import tqdm
from datetime import datetime
from conva_ai import AsyncConvaAI
from openpyxl import load_workbook
from conva_ai.response import ConvaAIResponse


nest_asyncio.apply()
current_date = datetime.now().strftime("%d")

def load_config(env, date):
    # with open('../polyglot/tools/testing/loadtest/MS_config_SDK_Python.yaml', 'r') as file:
    with open('/Users/divyac/Documents/Notebooks/MS_config_SDK_Python.yaml', 'r') as file:
        config = yaml.safe_load(file)
    
    if env in config['configurations']:
        env_config = config['configurations'][env]
        env_config['output_file_path'] = env_config['output_file_path'].replace("+current_date+", date)
        return env_config
    else:
        raise ValueError("Invalid environment specified")
    

def extract_properties(dict: ConvaAIResponse, no_parameter):
  message = input_query = reason = request_id = response_language = is_final = domain_name = app_name = category = llm_key = is_error = is_unsupported = tool_name = is_parameter_complete = ""
  list_parameters = related_queries = []
  try:
    # print(no_parameter)
    message = dict.message
    input_query = dict.input_query
    reason = dict.reason
    request_id = dict.request_id
    response_language = dict.response_language
    is_final = dict.is_final
    domain_name = dict.domain_name
    app_name = dict.app_name
    category = dict.category
    llm_key = dict.llm_key
    search_query = {param: dict.parameters.get(param, "") for param in parameters_value}
    list_parameters = search_query
    related_queries = dict.related_queries
    is_error = dict.is_error
    is_unsupported = dict.is_unsupported
    tool_name = dict.tool_name
    is_parameter_complete = dict.is_parameter_complete
     
    return request_id, input_query, message, reason, list_parameters, related_queries, response_language, is_final, domain_name, app_name, category, llm_key, is_error, is_unsupported, tool_name, is_parameter_complete 

  except json.JSONDecodeError:
    return "", "", "", "", [], [], "", "", "", "", "", "", "", "", "", ""


def generate_ouput(client, output_file_path, input_file_path, input_sheet_name):
    with open(output_file_path, "w") as outfile:
        wb = load_workbook(input_file_path)
        writer = csv.writer(outfile)

        if input_sheet_name in wb.sheetnames:
            print(input_sheet_name)
            sheet = wb[input_sheet_name]
            ws_index = wb.index(sheet)
            wb.active = ws_index
            ws = wb.active
            iter_rows = ws.iter_rows(min_row=2, values_only=True)
            writer.writerow(
                ["request_id", "input_query", "message", "reason"]+ parameters_value + ["related_queries", "response_language", "is_final", "domain_name", "app_name", "category", "llm_key", "is_error", "is_unsupported", "tool_name", "is_parameter_complete", "json"]
            )

        for row in iter_rows:
            sentence = row[3]
            print("Sentence:", sentence)
            final_response = asyncio.run(client.invoke_capability(sentence, stream=False, llm_key = "openai-gpt-4o-mini-2024-07-18"))
            print("final response:",final_response)
            request_id, input_query, message, reason, list_parameters, related_queries, response_language, is_final, domain_name, app_name, category, llm_key, is_error, is_unsupported, tool_name, is_parameter_complete = extract_properties(final_response, no_parameter)
            row = [request_id, input_query, message, reason] + [list_parameters[param] for param in parameters_value] + [related_queries, response_language, is_final, domain_name, app_name, category, llm_key, 
                    is_error, is_unsupported, tool_name, is_parameter_complete, final_response]
            writer.writerow(row)

if __name__ == "__main__":
    current_date = datetime.now().strftime("%d")
    parser = argparse.ArgumentParser(description='Load configuration based on the environment.')
    parser.add_argument('environment', type=str, help='The environment to load configuration for (stage_amazon, prod_amazon, dev_weechef)')
    
    args = parser.parse_args()

    try:
        config = load_config(args.environment, current_date)
        # Assigning output_file_path from the loaded configuration
        output_file_path = config['output_file_path']
        input_file_path = config['input_file_path']
        input_sheet_name = config['input_sheet_name']
        HOST = config['HOST']
        assis_id = config['assistant_id']
        API_KEY = config['API_KEY']
        assistant_version = config['assistant_version']
        parameters_value = config['para']
        no_parameter = len(parameters_value)
        print("output:", output_file_path)
        print("Input:", input_sheet_name)
        print("sheet:", output_file_path)
        print("Host:", HOST)
        print("Assis_id:", assis_id)
        print("API:", API_KEY)
        print("version:", assistant_version)
        client = AsyncConvaAI(assistant_id=assis_id, assistant_version=assistant_version, api_key=API_KEY, host=HOST)
        generate_ouput(client, output_file_path, input_file_path, input_sheet_name)
    except ValueError as e:
        print(e)
#  To run
#  /Users/divyac/Documents/Notebooks/my_env/bin/python /Users/divyac/Documents/Notebooks/MS_PySDK_File.py dev_freshdesk
