import csv
from datetime import datetime, timezone
import json
import base64
import requests
import hashlib
import uuid
import hmac
from openpyxl import load_workbook
from cryptography.fernet import Fernet
import sseclient
from tqdm import tqdm
from time import perf_counter_ns
import pandas as pd
import json
import csv
from sklearn.metrics import f1_score
import torch
from transformers import AutoTokenizer, AutoModel
from sklearn.metrics.pairwise import cosine_similarity
from sklearn.utils.multiclass import unique_labels
from tqdm import tqdm 
import numpy as np

input_file_path = "/Users/divyac/Downloads/omni_regression_test_suite (1).xlsx"
input_sheet_name = "Mylo_Divya"
input_csv_encoding = 'utf-8'
output_file_path = "/Users/divyac/Downloads/Mylo_json_split.csv"
output_csv_encoding = 'utf-8'
# HOST = "https://slang-omni-server-stage-xzevyhdsoq-el.a.run.app"
#HOST = "https://slang-omni-server-prod-xzevyhdsoq-el.a.run.app"
HOST = "https://omni-inference-stage.delightfulforest-40b0615c.centralindia.azurecontainerapps.io/"

#API_KEY = "1fb3a253f0f74ae387b1196308921aa7"  # stage tier api key
#API_KEY = "1a9eb148a92b476794f1b6e2739c9a5a"  # prod tier api key
API_KEY = "0c84fa340a1d44188cb0e399d9bb8a08"  # mylo prod tier api key

def generate_security_token(request_key, request_timestamp) -> str:
    salt = bytes("7j7efiQmVrCs43NA8wrzmcH2xCqWNoE6_C2rgqQSVTQ=".encode())
    fernet = Fernet(salt)
    return fernet.encrypt(f"{request_key}:{request_timestamp}".encode()).decode()


def generate_auth_hash(
    api_key, request_path, request_method, request_key, request_timestamp
) -> str:
    signing_key = base64.b64encode(
        f"{request_path}:{request_method}:{request_key}:{request_timestamp}".encode()
    )
    digest = hmac.new(signing_key, api_key.encode(), digestmod=hashlib.sha256).digest()
    computed_hmac = base64.b64encode(digest)
    return computed_hmac.decode()


def generate_auth_creds(
    api_key,
    request_path,
    request_method="POST",
) -> tuple[str, str]:
    request_key = uuid.uuid4().hex
    request_timestamp = (datetime.now(tz=timezone.utc)).timestamp()
    apikey_hash = generate_auth_hash(
        api_key, request_path, request_method, request_key, request_timestamp
    )
    token = generate_security_token(request_key, request_timestamp)
    return apikey_hash, token


def _create_request_path(assistant_id, request_type):
    return f"/v1/assistants/{assistant_id}/{request_type}"

def extract_properties(json_str):
  try:
    # Check if the input is a valid JSON string
    if isinstance(json_str, str):
      json_obj = json.loads(json_str)

      # # Debugging
      # print("Parameters:", json_obj.get("parameters"))
      # print("Filters:", json_obj.get("parameters", {}).get("filters"))

      app_name = json_obj.get("app_name", "")
      app_action = json_obj.get("app_action", "")
      category = json_obj.get("category", "")
      hints = json_obj.get("hints", [])
      message = json_obj.get("message", "")
      search_term = json_obj.get("parameters", {}).get("search_term", "")
      tool_name = json_obj.get("tool_name", "")
      filter_pairs = []

      # Extract filters (if they exist)
      if "parameters" in json_obj and "filters" in json_obj["parameters"]:
        filters = json_obj["parameters"]["filters"]
        for filter_info in filters:
          # Attempt to extract 'key' and 'value' from various locations:
          if "key" in filter_info and "parameters" in filter_info and "value" in filter_info["parameters"]:
            filter_key = filter_info["key"]
            filter_value = filter_info["parameters"]["value"]
          elif "filter_key" in filter_info and "filter_value" in filter_info:  # Check for alternative key names
            filter_key = filter_info["filter_key"]
            filter_value = filter_info["filter_value"]
          else:
            print("Warning: Unexpected filter format:", filter_info) 
            continue  # Skip to the next filter 
        
          filter_pairs.append((filter_key, filter_value))

      return app_name, app_action, category, hints, message, search_term, tool_name, filter_pairs
    else:
      return "", "", "", "", "", "", []
  except json.JSONDecodeError:
    return "", "", "", "", "", "", []

def _do_text2action_request(body, assistant_id):
    request_type = "text2action"
    uri = _create_request_path(assistant_id=assistant_id, request_type=request_type)
    formatted_url = f"{HOST}{uri}"
    apikey_hash, token = generate_auth_creds(
        API_KEY,
        uri,
    )
    request_headers = {
        "Authorization": apikey_hash,
        "X-Slang-Security-Token": token,
        "Content-Type": "application/json",
    }
    start_time = perf_counter_ns()  # Start the timer
    response = requests.post(
        formatted_url,
        stream=True,
        json=body,
        headers=request_headers,
    )
    client = sseclient.SSEClient(response)
    time_to_first_token = 0
    latency = None

    tmp = []
    for event in client.events():
        if time_to_first_token == 0:
            time_to_first_token = (perf_counter_ns() - start_time) / 1e6 # Time to first token
        tmp.append(event.data)

    latency = (perf_counter_ns() - start_time) / 1e6  # Total latency
    result = tmp[-1] if tmp else None
    return result, time_to_first_token, latency


with open(output_file_path, "w") as outfile:
    wb = load_workbook(input_file_path)
    writer = csv.writer(outfile)

    if input_sheet_name in wb.sheetnames:
        sheet = wb[input_sheet_name]
        ws_index = wb.index(sheet)
        wb.active = ws_index
        ws = wb.active
        iter_rows = ws.iter_rows(min_row=2, values_only=True)
        writer.writerow(
            ["request_id", "assistant_id", "query", "app_name", "app_action", "category", "hints",  "message", "search_term", "tool_name", "filter_pairs", "Time to first token (ms)", "Latency (ms)"]
        )
        for row in tqdm(iter_rows, desc="Processing rows", unit="row", dynamic_ncols=True):
            print(row[0])
            assistant_id = row[0]
            sentence = row[3]
            request_id = str(uuid.uuid4())
            expected_response = row[4]
            params = {
                "type": "text2action",  # text2Search
                "assistant_id": assistant_id,
                "assistant_version": "2.0.0", #6.0.4
                "request_id": request_id,
                "device_id": str(uuid.uuid4()),
                "tool_config": {
                    "multi_step_tool": {"llm": "azure-gpt-3.5-turbo"},
                    "AssistedShopping": {"llm": "azure-gpt-3.5-turbo"},
                    "single_step_tool": {"llm": "azure-gpt-3.5-turbo"},
                    "router": {"llm": "azure-gpt-3.5-turbo"}
                },
                "domain_name": "retail",
                "input_query": sentence,
                "app_context": {},
            }
            result, time_to_first_token, latency = _do_text2action_request(params, assistant_id)
            app_name, app_action, category, hints,  message, search_term, tool_name, filter_pairs = extract_properties(str(result))
            writer.writerow(
                [
                    request_id,
                    assistant_id,
                    sentence,
                    app_name, 
                    app_action, 
                    category, 
                    hints,  
                    message, 
                    search_term, 
                    tool_name, 
                    filter_pairs,
                    f"{time_to_first_token:.2f}" if time_to_first_token else "N/A",  # Convert to milliseconds
                    f"{latency:.2f}" if latency else "N/A",  # Convert to milliseconds
                ]
            )